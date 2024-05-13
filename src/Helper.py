import csv

import numpy as np
import openpyxl
from openpyxl.styles import PatternFill


def BinSearch(positions, position, width):
    """
    获取该位置在列表中的索引
    :param positions: 位置列表
    :param position: 某个位置
    :param width: 图像的宽度
    :return: 索引
    """
    l = 0
    r = len(positions)
    while l < r:
        mid = (l + r) // 2
        if positions[mid][0] * width + positions[mid][1] < position[0] * width + position[1]:
            l = mid + 1
        elif positions[mid][0] * width + positions[mid][1] > position[0] * width + position[1]:
            r = mid
        else:
            return mid
    return l


def SeqSearch(positions, position, width):
    for i in range(len(positions)):
        if positions[i] == position:
            return i


def mcie2000V(color1, color2):
    """
    据算两个颜色之间的颜色差
    :param color1: 颜色1
    :param color2: 颜色2
    :return: 颜色差(float)
    """
    L1, a1, b1 = color1
    L2, a2, b2 = color2
    avg_Lp = (L1 + L2) / 2.0

    C1 = np.sqrt(np.sum(np.power(color1[1:], 2)))
    C2 = np.sqrt(np.sum(np.power(color2[1:], 2)))

    avg_C1_C2 = (C1 + C2) / 2.0

    G = 0.5 * (1 - np.sqrt(np.power(avg_C1_C2, 7.0) / (np.power(avg_C1_C2, 7.0) + np.power(25.0, 7.0))))

    a1p = (1.0 + G) * a1
    a2p = (1.0 + G) * a2

    C1p = np.sqrt(np.power(a1p, 2.0) + np.power(b1, 2.0))
    C2p = np.sqrt(np.power(a2p, 2.0) + np.power(b2, 2.0))

    avg_C1p_C2p = (C1p + C2p) / 2.0

    h1p = np.degrees(np.arctan2(b1, a1p))
    h2p = np.degrees(np.arctan2(b2, a2p))
    h1p += (h1p < 0) * 360.0
    h2p += (h2p < 0) * 360.0

    avg_Hp = (((np.abs(h1p - h2p) > 180.0) * 360.0) + h1p + h2p) / 2.0

    T = 1 - 0.17 * np.cos(np.radians(avg_Hp - 30.0)) + \
        0.24 * np.cos(np.radians(2.0 * avg_Hp)) + \
        0.32 * np.cos(np.radians(3.0 * avg_Hp + 6.0)) - \
        0.20 * np.cos(np.radians(4.0 * avg_Hp - 63.0))

    diff_h2p_h1p = h2p - h1p
    delta_hp = diff_h2p_h1p + (np.abs(diff_h2p_h1p) > 180) * 360
    delta_hp -= (h2p > h1p) * 720.0

    delta_Lp = L2 - L1
    delta_Cp = C2p - C1p
    delta_Hp = 2.0 * np.sqrt(C1p * C2p) * np.sin(np.radians(delta_hp) / 2.0)

    S_L = 1 + ((0.015 * np.power(avg_Lp - 50.0, 2)) / np.sqrt(20.0 + np.power(avg_Lp - 50.0, 2)))
    S_C = 1 + 0.045 * avg_C1p_C2p
    S_H = 1 + 0.015 * avg_C1p_C2p * T

    delta_ro = 30.0 * np.exp(-(np.power((avg_Hp - 275.0) / 25.0, 2.0)))
    R_C = np.sqrt(np.power(avg_C1p_C2p, 7.0) / (np.power(avg_C1p_C2p, 7.0) + np.power(25.0, 7.0)))
    R_T = -2 * R_C * np.sin(2 * np.radians(delta_ro))

    return np.sqrt(np.power(delta_Lp / S_L, 2.0) + np.power(delta_Cp / S_C, 2.0) + np.power(delta_Hp / S_H, 2.0) + \
                   R_T * (delta_Cp / S_C) * (delta_Hp / S_H))


def mcie2000M(color1, color2s):
    """
    计算以一个颜色与色谱之间的颜色差，色谱的尺寸为n * 3: ndarray，
    都为标准Lab颜色，即：
    L: 0 ~ 100，
    a: -128 ~ 128
    b: -128 ~ 128
    :param color1: 颜色1
    :param color2s: 色谱
    :return: 颜色差数组
    """
    L, a, b = color1

    avg_Lp = (L + color2s[:, 0]) / 2.0

    C1 = np.sqrt(np.sum(np.power(color1[1:], 2)))
    C2 = np.sqrt(np.sum(np.power(color2s[:, 1:], 2), axis=1))

    avg_C1_C2 = (C1 + C2) / 2.0

    G = 0.5 * (1 - np.sqrt(np.power(avg_C1_C2, 7.0) / (np.power(avg_C1_C2, 7.0) + np.power(25.0, 7.0))))

    a1p = (1.0 + G) * a
    a2p = (1.0 + G) * color2s[:, 1]

    C1p = np.sqrt(np.power(a1p, 2) + np.power(b, 2))
    C2p = np.sqrt(np.power(a2p, 2) + np.power(color2s[:, 2], 2))

    avg_C1p_C2p = (C1p + C2p) / 2.0

    h1p = np.degrees(np.arctan2(b, a1p))
    h1p += (h1p < 0) * 360

    h2p = np.degrees(np.arctan2(color2s[:, 2], a2p))
    h2p += (h2p < 0) * 360

    avg_Hp = (((np.fabs(h1p - h2p) > 180) * 360) + h1p + h2p) / 2.0

    T = 1 - 0.17 * np.cos(np.radians(avg_Hp - 30)) + \
        0.24 * np.cos(np.radians(2 * avg_Hp)) + \
        0.32 * np.cos(np.radians(3 * avg_Hp + 6)) - \
        0.2 * np.cos(np.radians(4 * avg_Hp - 63))

    diff_h2p_h1p = h2p - h1p
    delta_hp = diff_h2p_h1p + (np.fabs(diff_h2p_h1p) > 180) * 360
    delta_hp -= (h2p > h1p) * 720

    delta_Lp = color2s[:, 0] - L
    delta_Cp = C2p - C1p
    delta_Hp = 2 * np.sqrt(C2p * C1p) * np.sin(np.radians(delta_hp) / 2.0)

    S_L = 1 + ((0.015 * np.power(avg_Lp - 50, 2)) / np.sqrt(20 + np.power(avg_Lp - 50, 2.0)))
    S_C = 1 + 0.045 * avg_C1p_C2p
    S_H = 1 + 0.015 * avg_C1p_C2p * T

    delta_ro = 30 * np.exp(-(np.power(((avg_Hp - 275) / 25), 2.0)))
    R_C = np.sqrt((np.power(avg_C1p_C2p, 7.0)) / (np.power(avg_C1p_C2p, 7.0) + np.power(25.0, 7.0)))
    R_T = -2 * R_C * np.sin(2 * np.radians(delta_ro))

    return np.sqrt(
        np.power(delta_Lp / S_L, 2) +
        np.power(delta_Cp / S_C, 2) +
        np.power(delta_Hp / S_H, 2) +
        R_T * (delta_Cp / S_C) * (delta_Hp / S_H))

def rgbToARGB(rgb):
    """
    将 RGB 十进制值转换为 ARGB 格式的十六进制值
    :param rgb: RGB 十进制值
    :return: ARGB 格式的十六进制值
    """
    # 将 RGB 值转换为 ARGB 格式的十六进制值
    argb = 'FF' + rgb[1:]
    rgbColor = rgb.split(',')
    # to hex str
    for i in range(3):
        rgbColor[i] = hex(int(rgbColor[i]))[2:].upper()
        if len(rgbColor[i]) == 1:
            rgbColor[i] = '0' + rgbColor[i]
    argb = 'FF' + ''.join(rgbColor)  # 在 RGB 值前加上 FF 表示完全不透明
    return argb


def exportToExcel(colorsDict, filename):
    """
    将颜色字典导出到 Excel 文件
    :param colorsDict: 颜色字典
    :param filename: 文件名
    :return:
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet.cell(row=1, column=1, value='颜色')
    sheet.cell(row=1, column=2, value='值')
    sheet.cell(row=1, column=3, value='数量')

    row = 2
    for color, value in colorsDict.items():
        argb = rgbToARGB(color)  # 将 RGB 值转换为 ARGB 格式
        fill = PatternFill(fill_type='solid', fgColor=argb)
        cell = sheet.cell(row=row, column=1)
        cell.fill = fill
        sheet.cell(row=row, column=2, value=f'rgb({color})')
        sheet.cell(row=row, column=3, value=value[0])
        row += 1
    workbook.save(filename)


def exportToCsv(colorDict, filename):
    """
    将颜色字典导出到 CSV 文件
    :param colorDict: 颜色字典
    :param filename: 文件名
    :return:
    """
    with open(filename, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(['值', '数量'])
        for color, value in colorDict.items():
            writer.writerow([f'({color})', value[0]])


if __name__ == '__main__':
    c1 = [1, 0, 1]
    c2 = [2, 1, 3]
    res = mcie2000V(c1, c2)
    print(res)
